"""
PDF, CSV, and Excel export utilities for BauDok reports.
"""

import io
from django.template.loader import render_to_string


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def generate_pdf(report) -> bytes:
    """Render DailyReport to PDF bytes using WeasyPrint."""
    from weasyprint import HTML

    context = {
        'report': report,
        'entries': report.entries.all(),
        'company': report.company,
        'signatures': report.signatures.all(),
    }
    html_content = render_to_string('reports/daily_report.html', context)
    return HTML(string=html_content).write_pdf()


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def generate_csv(reports) -> bytes:
    """Generate a CSV export for a list of DailyReport instances using pandas."""
    import pandas as pd

    rows = []
    for report in reports:
        base = {
            'Datum': report.report_date,
            'Projekt': report.project.name if report.project else '',
            'Mitarbeiter': report.created_by.get_full_name() if report.created_by else '',
            'Gewerk': report.created_by.trade if report.created_by else '',
            'Status': report.get_status_display(),
            'Wetter': report.weather,
            'Temperatur': report.temperature,
            'KI-Modell': report.ai_model_used,
            'Tokens': report.ai_tokens_used,
            'Erstellt': report.created_at.strftime('%Y-%m-%d %H:%M') if report.created_at else '',
        }
        for entry in report.entries.all():
            row = base.copy()
            row['Kategorie'] = entry.get_category_display()
            row['Inhalt'] = entry.content
            row['Dauer (h)'] = entry.duration_hours
            row['Menge'] = entry.quantity
            rows.append(row)

        if not report.entries.exists():
            rows.append(base)

    df = pd.DataFrame(rows)
    buffer = io.BytesIO()
    df.to_csv(buffer, index=False, encoding='utf-8-sig')
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def generate_excel(reports) -> bytes:
    """Generate an Excel export for a list of DailyReport instances using openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tagesberichte'

    headers = [
        'Datum', 'Projekt', 'Mitarbeiter', 'Gewerk', 'Status',
        'Wetter', 'Temperatur', 'Kategorie', 'Inhalt', 'Dauer (h)', 'Menge',
        'KI-Modell', 'Tokens', 'Erstellt',
    ]

    # Header row styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.row_dimensions[1].height = 25

    row_idx = 2
    for report in reports:
        base = [
            report.report_date,
            report.project.name if report.project else '',
            report.created_by.get_full_name() if report.created_by else '',
            report.created_by.trade if report.created_by else '',
            report.get_status_display(),
            report.weather,
            report.temperature,
        ]

        entries = list(report.entries.all())
        if entries:
            for entry in entries:
                row_data = base + [
                    entry.get_category_display(),
                    entry.content,
                    float(entry.duration_hours) if entry.duration_hours else None,
                    entry.quantity,
                    report.ai_model_used,
                    report.ai_tokens_used,
                    report.created_at.strftime('%Y-%m-%d %H:%M') if report.created_at else '',
                ]
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                row_idx += 1
        else:
            row_data = base + ['', '', None, '', report.ai_model_used, report.ai_tokens_used,
                                report.created_at.strftime('%Y-%m-%d %H:%M') if report.created_at else '']
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

    # Auto-fit column widths (approximate)
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
